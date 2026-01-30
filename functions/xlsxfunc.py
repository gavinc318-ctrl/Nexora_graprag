"""code/functions/xlsxfunc.py

XLSX (Excel) -> logical pages for embedding-based retrieval.

Design goals
- Stable, reproducible logical pagination (no Excel rendering dependencies).
- Preserve workbook structure via sheet boundaries.
- Generate embedding-friendly text: sheet summary + row narratives.
- Keep a compact table snapshot and formula evidence for traceability.

Public API (do NOT change signature):
    parse_xlsx_to_pages(file_path: str) -> List[Dict[str, Any]]

Return format (matches TXT/DOCX parsers):
    [
      {"page_no": 1, "text": "...", "tables": "...", "png_bytes": None},
      ...
    ]

Notes
- We read workbook twice:
  * data_only=True: prefers cached formula results for recall text.
  * data_only=False: captures original formulas for evidence.
- If formula results are not cached (common for files not saved after calc), the
  value may be None. We keep the row narrative but leave the field blank.
"""

from __future__ import annotations

import datetime as _dt
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# -----------------------------
# Small helpers
# -----------------------------

def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _norm_space(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def _format_value(v: Any) -> str:
    """User-friendly, embedding-friendly cell value formatting."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        # ISO-like format
        if isinstance(v, _dt.datetime):
            return v.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(v, _dt.date):
            return v.strftime("%Y-%m-%d")
        return v.strftime("%H:%M:%S")
    if isinstance(v, (int, float)):
        # Keep compact but readable (avoid scientific notation surprises)
        if isinstance(v, float):
            if abs(v) >= 1e10 or (0 < abs(v) < 1e-6):
                return f"{v:.6g}"
            return f"{v:.6f}".rstrip("0").rstrip(".")
        return str(v)
    return _norm_space(str(v))


def _trim_used_range(values_grid: List[List[Any]]) -> Tuple[List[List[Any]], int, int]:
    """Trim trailing empty rows/cols. Returns (trimmed_grid, nrows, ncols)."""
    if not values_grid:
        return [], 0, 0

    # Trim empty trailing rows
    last_row = len(values_grid)
    while last_row > 0 and all(_is_empty(x) for x in values_grid[last_row - 1]):
        last_row -= 1
    values_grid = values_grid[:last_row]

    if not values_grid:
        return [], 0, 0

    # Determine last non-empty column across remaining rows
    last_col = 0
    for row in values_grid:
        for j in range(len(row) - 1, -1, -1):
            if not _is_empty(row[j]):
                last_col = max(last_col, j + 1)
                break

    values_grid = [r[:last_col] for r in values_grid]

    # Trim empty leading rows? We keep leading blanks for now; header detection will skip.
    return values_grid, len(values_grid), last_col


def _find_header_row(grid: List[List[Any]]) -> Optional[int]:
    """Return 0-based header row index; heuristic: first row with >=2 non-empty cells."""
    for i, row in enumerate(grid):
        non_empty = sum(0 if _is_empty(x) else 1 for x in row)
        if non_empty >= 2:
            return i
    # If everything is sparse, accept first non-empty row
    for i, row in enumerate(grid):
        if any(not _is_empty(x) for x in row):
            return i
    return None


_DIM_HINTS = [
    "date", "day", "month", "year", "time", "week", "region", "area", "country",
    "province", "city", "state", "product", "sku", "item", "category", "type",
    "name", "id", "code", "channel", "dept", "department", "customer", "vendor",
    "segment", "group",
    # Chinese common
    "日期", "时间", "月份", "年度", "地区", "区域", "国家", "省", "市", "产品", "品类",
    "名称", "编号", "代码", "渠道", "部门", "客户", "供应商", "分组",
]

_METRIC_HINTS = [
    "amount", "sales", "revenue", "income", "cost", "expense", "profit", "margin",
    "rate", "%", "qty", "quantity", "count", "total", "avg", "average", "score",
    "price", "value",
    # Chinese common
    "金额", "销售", "收入", "成本", "费用", "利润", "毛利", "毛利率", "同比", "环比",
    "增长", "数量", "合计", "均值", "单价", "得分", "比率", "率", "占比",
]


def _classify_columns(headers: Sequence[str]) -> Tuple[List[str], List[str]]:
    dims: List[str] = []
    metrics: List[str] = []
    for h in headers:
        hl = h.lower()
        if any(k in hl for k in _METRIC_HINTS):
            metrics.append(h)
        elif any(k in hl for k in _DIM_HINTS):
            dims.append(h)
    return dims, metrics


def _build_table_snapshot(headers: List[str], rows: List[List[Any]], max_rows: int = 50) -> str:
    """Create a compact TSV table snapshot for audit/debug."""
    lines: List[str] = []
    if headers:
        lines.append("\t".join(_norm_space(str(h)) for h in headers))
    take = rows[:max_rows]
    for r in take:
        lines.append("\t".join(_format_value(x) for x in r))
    if len(rows) > max_rows:
        lines.append(f"[TRUNCATED] showing first {max_rows} rows of {len(rows)}")
    return "\n".join(lines).strip()


def _extract_formulas(ws_formula, max_row: int, max_col: int, max_items: int = 200) -> List[Tuple[str, str]]:
    """Return list of (A1, formula_str)."""
    out: List[Tuple[str, str]] = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws_formula.cell(row=r, column=c)
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                addr = f"{get_column_letter(c)}{r}"
                out.append((addr, v))
                if len(out) >= max_items:
                    return out
    return out


# -----------------------------
# Public API
# -----------------------------

def parse_xlsx_to_pages(
    file_path: str,
    rows_per_page: int = 200,
    snapshot_rows: int = 50,
) -> List[Dict[str, Any]]:
    """Parse XLSX into logical pages.

    Pagination strategy
    - Iterate sheets in workbook order.
    - Each sheet yields 1+ pages; split by `rows_per_page` for very large sheets.

    Each page's `text` is embedding-focused:
    - a sheet/part summary
    - followed by row narratives "Header=Value".

    `tables` contains a compact TSV snapshot (first `snapshot_rows` rows) and a
    capped list of formulas for traceability.
    """

    p = Path(file_path)
    if not p.exists() or not p.is_file():
        raise FileNotFoundError(f"xlsx file not found: {file_path}")
    if p.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        # core may route .xls here too; .xls is not supported by openpyxl.
        raise ValueError(
            f"unsupported Excel format for openpyxl: {p.suffix}. "
            "Please convert .xls to .xlsx first."
        )

    # Values workbook (prefers cached results of formulas)
    wb_val = load_workbook(filename=str(p), data_only=True, read_only=True)
    # Formula workbook (captures original formulas)
    wb_for = load_workbook(filename=str(p), data_only=False, read_only=True)

    pages: List[Dict[str, Any]] = []
    page_no = 1

    for sheet_name in wb_val.sheetnames:
        ws_val = wb_val[sheet_name]
        ws_for = wb_for[sheet_name]

        # Build a raw grid from the worksheet's current dimensions.
        # read_only worksheets still support .max_row/.max_column.
        max_row = int(ws_val.max_row or 0)
        max_col = int(ws_val.max_column or 0)
        if max_row <= 0 or max_col <= 0:
            # Empty sheet -> still emit one page to keep structure
            text = f"[[META type=excel sheet={sheet_name} part=1/1]]\nSheet: {sheet_name}\n[EMPTY SHEET]"
            pages.append({"page_no": page_no, "text": text, "tables": "", "png_bytes": None})
            page_no += 1
            continue

        grid: List[List[Any]] = []
        for r in range(1, max_row + 1):
            row_vals: List[Any] = []
            for c in range(1, max_col + 1):
                row_vals.append(ws_val.cell(row=r, column=c).value)
            grid.append(row_vals)

        grid, nrows, ncols = _trim_used_range(grid)
        if nrows == 0 or ncols == 0:
            text = f"[[META type=excel sheet={sheet_name} part=1/1]]\nSheet: {sheet_name}\n[EMPTY SHEET]"
            pages.append({"page_no": page_no, "text": text, "tables": "", "png_bytes": None})
            page_no += 1
            continue

        header_idx = _find_header_row(grid)
        if header_idx is None:
            header_idx = 0

        raw_headers = grid[header_idx]
        headers: List[str] = []
        for j, h in enumerate(raw_headers, start=1):
            hv = _format_value(h)
            if hv:
                headers.append(hv)
            else:
                headers.append(f"{get_column_letter(j)}")

        data_rows = grid[header_idx + 1 :]

        # Build table snapshot rows: include header + first N data rows.
        snapshot = _build_table_snapshot(headers, data_rows, max_rows=snapshot_rows)

        # Formula evidence from the *formula* workbook (scan only within trimmed cols/rows)
        formulas = _extract_formulas(ws_for, max_row=nrows, max_col=ncols, max_items=200)
        formula_block = ""
        if formulas:
            formula_lines = ["[FORMULAS] (capped)"]
            for addr, f in formulas:
                formula_lines.append(f"{addr}\t{f}")
            if len(formulas) >= 200:
                formula_lines.append("[TRUNCATED] showing first 200 formulas")
            formula_block = "\n".join(formula_lines)

        tables_out = "\n\n".join(x for x in [snapshot, formula_block] if x.strip()).strip()

        dims, metrics = _classify_columns(headers)

        # Split into pages by rows_per_page
        if rows_per_page <= 0:
            rows_per_page = 200

        total_rows = len(data_rows)
        total_parts = max(1, (total_rows + rows_per_page - 1) // rows_per_page)

        for part_idx in range(total_parts):
            start = part_idx * rows_per_page
            end = min(total_rows, (part_idx + 1) * rows_per_page)
            part_rows = data_rows[start:end]

            # Build embedding-focused text
            meta = f"[[META type=excel sheet={sheet_name} part={part_idx+1}/{total_parts}]]"
            summary_lines: List[str] = [
                meta,
                f"Sheet: {sheet_name}",
                f"UsedRange: rows={header_idx+1}-{header_idx+1+total_rows}, cols=1-{ncols}",
                f"Columns: {', '.join(headers)}",
            ]
            if dims:
                summary_lines.append(f"Likely dimensions: {', '.join(dims)}")
            if metrics:
                summary_lines.append(f"Likely metrics: {', '.join(metrics)}")

            if total_parts > 1:
                summary_lines.append(f"Rows in this part: {start+1}-{end} of {total_rows}")

            summary_text = "\n".join(summary_lines).strip()

            row_lines: List[str] = []
            # We use original Excel row numbers for traceability: header row is (header_idx+1)
            excel_row_no = header_idx + 2 + start
            for row in part_rows:
                pairs: List[str] = []
                for h, v in zip(headers, row):
                    fv = _format_value(v)
                    if fv == "":
                        continue
                    pairs.append(f"{h}={fv}")
                if not pairs:
                    excel_row_no += 1
                    continue
                row_lines.append(f"Row {excel_row_no}: " + ", ".join(pairs))
                excel_row_no += 1

            text_out = (summary_text + "\n\n" + "\n".join(row_lines)).strip()

            pages.append(
                {
                    "page_no": page_no,
                    "text": text_out,
                    "tables": tables_out,
                    "png_bytes": None,
                }
            )
            page_no += 1

    # Safety: ensure non-empty pages list
    if not pages:
        pages = [{"page_no": 1, "text": "", "tables": "", "png_bytes": None}]

    return pages
